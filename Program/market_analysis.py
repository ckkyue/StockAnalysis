# Import
import datetime as dt
from helper_functions import get_current_date, get_df, get_excel_filename, stock_market
import openpyxl
from openpyxl.styles import Font, PatternFill
from pandas import ExcelWriter as EW
from plot import *
import scipy.stats as stats
from technicals import *

# # Display the whole dataframe
# pd.set_option("display.max_rows", None)
# pd.set_option("display.max_columns", None)

# Screen the stocks from Excel file
def screen_excel(excel_filename, sectors_excel_leading, sectors_excel_improving):
    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(excel_filename)
    sheet = workbook.active

    # Define the fill colour for highlighting
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    red_font = Font(color="FF0000")
    green_font = Font(color="008000")

    # Find the index of the "Stock", "Sector", "Volatility 20 Z-Score", "Volatility 60 Z-Score", and "VCP" columns
    stock_col_index = None
    volatility20_col_index = None
    volatility60_col_index = None
    mvp_col_index = None
    vcp_col_index = None
    sector_col_index = None
    
    for cell in sheet[1]: # Assuming the first row contains headers
        if cell.value == "Stock":
            stock_col_index = cell.column
        elif cell.value == "Volatility 20 Z-Score":
            volatility20_col_index = cell.column
        elif cell.value == "Volatility 60 Z-Score":
            volatility60_col_index = cell.column
        elif cell.value == "MVP":
            mvp_col_index = cell.column
        elif cell.value == "VCP":
            vcp_col_index = cell.column
        elif cell.value == "Sector":
            sector_col_index = cell.column

    # Highlight the cells of each row
    if sector_col_index is not None:
        for row in sheet.iter_rows(min_row=2): # Start from the second row
            stock_cell = row[stock_col_index - 1] # Adjust for zero-based index
            volatility20_cell = row[volatility20_col_index - 1]
            volatility60_cell = row[volatility60_col_index - 1]
            sector_cell = row[sector_col_index - 1]
            mvp_cell = row[mvp_col_index - 1]
            vcp_cell = row[vcp_col_index - 1]
            
            # Change text colour to red if Volatility 20 Z-Score is greater than 2
            if volatility20_cell.value > 2:
                volatility20_cell.font = red_font
            elif volatility20_cell.value < - 1:
                volatility20_cell.font = green_font

            # Change text colour to red if Volatility 60 Z-Score is greater than 2
            if volatility60_cell.value > 2:
                volatility60_cell.font = red_font
            elif volatility60_cell.value < - 1:
                volatility60_cell.font = green_font

            # Highlight the stock if its sector matches
            if sector_cell.value in sectors_excel_leading + sectors_excel_improving:
                stock_cell.fill = yellow_fill
                sector_cell.fill = yellow_fill

            # Change text colour to green if MVP
            if mvp_cell.value == "MVP":
                mvp_cell.font = green_font

            # Highlight VCP
            if vcp_cell.value == True:
                vcp_cell.fill = orange_fill

    # Save the changes to the Excel file
    workbook.save(excel_filename)
    print(f"Changes made to the Excel file {excel_filename}.")

# Calculate the retracement for stocks
def retracement_excel(excel_filename, end_date, min_column="Low", max_column="High", period=5, buffer=15):
    # Read the screened stocks
    df = pd.read_excel(excel_filename)

    # Initialize lists to store retracement values and SMA 5 slopes
    retracements = []
    SMA_5_slopes = []

    # Get the list of stocks
    stocks = df["Stock"].tolist()

    # Iterate over all stocks
    for stock in stocks:
        data = get_df(stock, end_date)
        data = get_local_extrema(data, min_column=min_column, max_column=max_column, period=period)
        data["Percent Change"] = data["Close"].pct_change()
        data["Percent Change SMA 5"] = SMA(data, 5, "Percent Change")
        SMA_5_slope = data["Percent Change SMA 5"].iloc[-1]
        SMA_5_slopes.append(SMA_5_slope)
        local_min1, local_max, retracement = calculate_retracement(data, min_column=min_column, max_column=max_column, buffer=buffer)
        retracements.append(retracement)

    SMA_5_slopes = np.array(SMA_5_slopes)
    SMA_5_slopes_zscore = stats.zscore(SMA_5_slopes)
    retracements = np.array(retracements)
    retracements_zscore = stats.zscore(retracements)

    # Overwrite or insert the SMA 5 slopes, retracement values, and their z-scores
    if "SMA 5 Slope (%)" in df.columns:
        df["SMA 5 Slope (%)"] = np.round(SMA_5_slopes * 100, 2)
    else:
        df.insert(df.columns.get_loc("Close") + 1, "SMA 5 Slope (%)", np.round(SMA_5_slopes * 100, 2))

    if "SMA 5 Slope Z-Score" in df.columns:
        df["SMA 5 Slope Z-Score"] = np.round(SMA_5_slopes * 100, 2)
    else:
        df.insert(df.columns.get_loc("SMA 5 Slope (%)") + 1, "SMA 5 Slope Z-Score", SMA_5_slopes_zscore)
    
    if "Retracement (%)" in df.columns:
        df["Retracement (%)"] = np.round(retracements * 100, 2)
    else:
        df.insert(df.columns.get_loc("SMA 5 Slope (%)") + 1, "Retracement (%)", np.round(retracements * 100, 2))

    if "Retracement Z-Score" in df.columns:
        df["Retracement Z-Score"] = retracements_zscore
    else:
        df.insert(df.columns.get_loc("Retracement (%)") + 1, "Retracement Z-Score", retracements_zscore)

    # Save the changes to the Excel file
    writer = EW(excel_filename)
    df.to_excel(writer, sheet_name="Sheet1", index=False)
    writer._save()
    print(f"Changes made to the Excel file {excel_filename}.")

# Main function
def main():
    # Start of the program
    start = dt.datetime.now()

    # Initial setup
    current_date = get_current_date(start)

   # Variables
    NASDAQ_all = True
    period_hk = 60 # Period for HK stocks
    period_us = 252 # Period for US stocks
    RS = 90

    # Index
    index_name = "^GSPC"
    index_names = ["^HSI", "^GSPC", "^IXIC", "^DJI", "IWM", "FFTY", "QQQE", "GC=F"]
    index_dict = {"^HSI": "HKEX", "^GSPC": "S&P 500", "^IXIC": "NASDAQ Composite", "^DJI": "Dow Jones Industrial Average", 
                  "IWM": "iShares Russell 2000 ETF", "FFTY": "Innovator IBD 50 ETF", "QQQE": "NASDAQ-100 Equal Weighted ETF", 
                  "KWEB": "KraneShares CSI China Internet ETF", "GC=F": "Gold"}
     
    # Sectors
    sectors = ["XLC", "XLY", "XLP", "XLE", "XLF", "XLV", 
            "XLI", "XLB", "XLRE", "XLK", "XLU"]
    sector_dict = {"XLC": "Communication Services", "XLY": "Consumer Discretionary", "XLP": "Consumer Staples", 
                "XLE": "Energy", "XLF": "Financials", "XLV": "Health Care", 
                "XLI": "Industrials", "XLB": "Materials", "XLRE": "Real Estate", 
                "XLK": "Technology", "XLU": "Utilities"}
    sector_excel_dict = {"XLC": "Communication Services", "XLY": "Consumer Cyclical", "XLP": "Consumer Defensive", 
                "XLE": "Energy", "XLF": "Financial Services", "XLV": "Healthcare", 
                "XLI": "Industrials", "XLB": "Basic Materials", "XLRE": "Real Estate", 
                "XLK": "Technology", "XLU": "Utilities"}
    
    # Define the result folder
    result_folder = "Result"
    
    # Get the price data of the index
    index_df = get_df(index_name, current_date)

    plot_all = False
    if plot_all:
        # Iterate over all indices and sectors
        for ticker in index_names + sectors:
            # Get the price data of the tickers
            df = get_df(ticker, current_date)

            # Visualize the closing price history of the ticker
            plot_close(ticker, df, MVP_VCP=False, save=True)

    sector_rotation = True
    if sector_rotation:
        # Calculate the JdK RS-Ratio and Momentum
        index_df = get_JdK(index_names + sectors, index_df, current_date)

        # Print the leading, weakening, improving and lagging sectors
        sectors_leading = []
        sectors_excel_leading = []
        sectors_weakening = []
        sectors_improving = []
        sectors_excel_improving = []
        sectors_lagging = []

        # Iterate over all sectors
        for sector in sectors:
            if index_df[f"{sector} JdK RS-Ratio"].iloc[-1] > 100:
                if index_df[f"{sector} JdK RS-Momentum"].iloc[-1] > 100:
                    sectors_leading.append(sector_dict[sector])
                    sectors_excel_leading.append(sector_excel_dict[sector])
                else:
                    sectors_weakening.append(sector_dict[sector])
            else:
                if index_df[f"{sector} JdK RS-Momentum"].iloc[-1] > 100:
                    sectors_improving.append(sector_dict[sector])
                    sectors_excel_improving.append(sector_excel_dict[sector])
                else:
                    sectors_lagging.append(sector_dict[sector])

        # Print the classified sectors
        print(f"Leading sectors: {', '.join(sectors_leading)}")
        print(f"Weakening sectors: {', '.join(sectors_weakening)}")
        print(f"Improving sectors: {', '.join(sectors_improving)}")
        print(f"Lagging sectors: {', '.join(sectors_lagging)}")

    plot_all_jdk = False
    if plot_all_jdk:
        # Iterate over all sectors
        for sector in sectors:
            # Plot the JdK RS-Ratio and Momentum of the sector
            plot_JdK(sector, sector_dict, index_df, save=True)

    sector_selected = False
    if sector_selected:
        # Plot the relative rotation graph
        plot_rrg(sectors, sector_dict, index_df, "sector", save=True)
        
        # Plot the sectors of the selected stocks
        plot_sector_selected(current_date, "^GSPC", index_dict, NASDAQ_all=NASDAQ_all, save=True)
    
    hkex_retracement = False
    if hkex_retracement:
        # Get the Excel filename
        excel_filename = get_excel_filename(current_date, "^HSI", index_dict, period_hk, period_us, RS, NASDAQ_all, result_folder)

        retracement_excel(excel_filename, current_date)

    screen = True
    if screen:
        # Get the Excel filename
        excel_filename = get_excel_filename(current_date, "^GSPC", index_dict, period_hk, period_us, RS, NASDAQ_all, result_folder)

        # Screen the stocks from Excel file
        screen_excel(excel_filename, sectors_excel_leading, sectors_excel_improving)

        # Get the Excel filename
        excel_filename = get_excel_filename(current_date, "^HSI", index_dict, period_hk, period_us, RS, NASDAQ_all, result_folder)

        # Screen the stocks from Excel file
        screen_excel(excel_filename, sectors_excel_leading, sectors_excel_improving)

    plot_marketbreadth = False
    if plot_marketbreadth:
        # Get the list of tickers of stock market
        index_df = get_df(index_name, current_date)
        tickers = stock_market(current_date, current_date, index_name, False)

        # Calculate the market breadth indicators
        index_df = market_breadth(current_date, index_df, tickers)

        # Save the data of the index to a .csv file
        index_filename = f"Price data/{index_name}_{current_date}.csv"
        index_df.to_csv(index_filename)

        # Visualize the closing price history and other technical indicators
        plot_market_breadth(index_name, index_df, tickers, save=True)
        plot_close(index_name, index_df, MVP_VCP=False)
        plot_MFI_RSI(index_name, index_df, 252, save=True)
    
    plot_vix = False
    if plot_vix:
        # Get the price data of CBOE Volatility Index (VIX)
        vix_df = get_df("^VIX", current_date)

        # Get the current closing price of VIX
        vix_current_high = round(vix_df["High"].iloc[-1], 2)

        # Define the exit indicator based on VIX value
        if vix_current_high < 26:
            vix_colour = "Green"
        elif 26 < vix_current_high < 30:
            vix_colour = "Yellow"
        elif vix_current_high > 30:
            vix_colour = "Red"

        # Plot the closing price history of VIX
        plot_close("^VIX", vix_df, save=True)

        # Print the exit indicator
        print(f"Current VIX: {vix_current_high} ({vix_colour})")

    # Print the end time and total runtime
    end = dt.datetime.now()
    print(end, "\n")
    print("The program used", end - start)

# Run the main function
if __name__ == "__main__":
    main()